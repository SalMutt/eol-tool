"""Tests for the input filter module."""

from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import openpyxl
from click.testing import CliRunner

from eol_tool.checker import BaseChecker
from eol_tool.cli import cli
from eol_tool.input_filter import filter_models, is_junk_row
from eol_tool.models import EOLResult, EOLStatus, HardwareModel

# ---------------------------------------------------------------------------
# is_junk_row — junk rows (should be filtered)
# ---------------------------------------------------------------------------


class TestJunkRows:
    def test_inventory_code_no_manufacturer(self):
        assert is_junk_row("UK-4216", "") is True

    def test_server_barebone_no_manufacturer(self):
        assert is_junk_row("SERVER BAREBONE", "") is True

    def test_ram_config_no_manufacturer(self):
        assert is_junk_row("2 X 4316 - 768GB", "") is True

    def test_capacity_speed_no_manufacturer(self):
        assert is_junk_row("1TBI 7K", "") is True

    def test_cpu_series_no_manufacturer(self):
        assert is_junk_row("E3-1200 SERIES", "") is True

    def test_half_slim_ssd_no_manufacturer(self):
        assert is_junk_row("HALF-SLIM SSD", "") is True

    def test_channel_range_no_manufacturer(self):
        assert is_junk_row("4CH 21-24", "") is True

    def test_new_no_manufacturer(self):
        assert is_junk_row("NEW", "") is True

    def test_used_no_manufacturer(self):
        assert is_junk_row("USED", "") is True

    def test_rw_server_no_manufacturer(self):
        assert is_junk_row("RW SERVER UK", "") is True

    def test_tbram_no_manufacturer(self):
        assert is_junk_row("7960 1TBRAM", "") is True

    def test_ams_server_no_manufacturer(self):
        assert is_junk_row("AMS1 SERVER", "") is True

    def test_fs_box_no_manufacturer(self):
        assert is_junk_row("FS BOX", "") is True

    def test_empty_model_no_manufacturer(self):
        assert is_junk_row("", "") is True

    def test_whitespace_model_no_manufacturer(self):
        assert is_junk_row("   ", "") is True


# ---------------------------------------------------------------------------
# is_junk_row — kept rows (should NOT be filtered)
# ---------------------------------------------------------------------------


class TestKeptRows:
    def test_optics_qsfp_no_manufacturer(self):
        assert is_junk_row("QSFP-SR4-40G", "") is False

    def test_optics_sfp_no_manufacturer(self):
        assert is_junk_row("SFP-10G-T", "") is False

    def test_optics_sfp_plus_no_manufacturer(self):
        assert is_junk_row("10G SFP+ LR", "") is False

    def test_real_hardware_with_manufacturer(self):
        assert is_junk_row("ASR-6405E", "Adaptec") is False

    def test_drive_with_manufacturer(self):
        assert is_junk_row("0F12470", "Hitachi") is False

    def test_manufacturer_always_keeps(self):
        assert is_junk_row("SERVER BAREBONE", "Dell") is False
        assert is_junk_row("NEW", "Intel") is False
        assert is_junk_row("UK-4216", "SomeVendor") is False

    def test_optics_xfp_no_manufacturer(self):
        assert is_junk_row("XFP-10G-SR", "") is False

    def test_optics_cfp_no_manufacturer(self):
        assert is_junk_row("CFP-100G-LR4", "") is False

    def test_dwdm_sfpp_c30_no_manufacturer(self):
        assert is_junk_row("C30 SFPP-10G-DW30", "") is False

    def test_dwdm_sfpp_c31_no_manufacturer(self):
        assert is_junk_row("C31 SFPP-10G-DW31", "") is False

    def test_dwdm_sfpp_c32_no_manufacturer(self):
        assert is_junk_row("C32 SFPP-10G-DW32", "") is False

    def test_optics_sfp1g_no_manufacturer(self):
        assert is_junk_row("SFP1G-LX-31", "") is False


# ---------------------------------------------------------------------------
# filter_models
# ---------------------------------------------------------------------------


class TestFilterModels:
    def test_splits_clean_and_filtered(self):
        models = [
            HardwareModel(model="QSFP-SR4-40G", manufacturer="", category="optics"),
            HardwareModel(model="UK-4216", manufacturer="", category="unknown"),
            HardwareModel(model="EX4300-48T", manufacturer="Juniper", category="switch"),
            HardwareModel(model="NEW", manufacturer="", category="unknown"),
        ]
        clean, filtered = filter_models(models)
        assert len(clean) == 2
        assert len(filtered) == 2
        assert clean[0].model == "QSFP-SR4-40G"
        assert clean[1].model == "EX4300-48T"

    def test_filtered_rows_contain_reason(self):
        models = [
            HardwareModel(model="UK-4216", manufacturer="", category="unknown"),
        ]
        _, filtered = filter_models(models)
        assert len(filtered) == 1
        assert filtered[0]["model"] == "UK-4216"
        assert filtered[0]["manufacturer"] == ""
        assert filtered[0]["reason"]

    def test_all_clean_returns_empty_filtered(self):
        models = [
            HardwareModel(model="EX4300-48T", manufacturer="Juniper", category="switch"),
        ]
        clean, filtered = filter_models(models)
        assert len(clean) == 1
        assert len(filtered) == 0

    def test_all_junk_returns_empty_clean(self):
        models = [
            HardwareModel(model="NEW", manufacturer="", category="unknown"),
            HardwareModel(model="USED", manufacturer="", category="unknown"),
        ]
        clean, filtered = filter_models(models)
        assert len(clean) == 0
        assert len(filtered) == 2


# ---------------------------------------------------------------------------
# Filtered sheet in output xlsx
# ---------------------------------------------------------------------------


class _FakeChecker(BaseChecker):
    manufacturer_name = "testmfr"
    rate_limit = 10

    async def check(self, model: HardwareModel) -> EOLResult:
        return EOLResult(
            model=model,
            status=EOLStatus.ACTIVE,
            checked_at=datetime.now(),
            source_name="fake",
            confidence=90,
        )


def _write_test_xlsx(path: Path, rows: list[tuple[str, str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Models"
    ws.append(["Model", "Manufacturer", "Category", "Condition", "Original_Item"])
    for model, mfr, cat in rows:
        ws.append([model, mfr, cat, "", ""])
    wb.save(path)
    wb.close()


class TestFilteredXlsxSheet:
    def test_filtered_sheet_in_output(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_test_xlsx(
            input_path,
            [
                ("EX4300-48T", "Juniper", "switch"),
                ("NEW", "", "unknown"),
                ("UK-4216", "", "unknown"),
            ],
        )
        output_path = tmp_path / "output.xlsx"
        runner = CliRunner()

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
        ):
            mock_list.return_value = {"__fallback__": _FakeChecker}
            mock_get.side_effect = (
                lambda name: _FakeChecker if name == "__fallback__" else None
            )
            result = runner.invoke(
                cli,
                [
                    "check",
                    "--input", str(input_path),
                    "--output", str(output_path),
                    "--no-cache",
                ],
            )

        assert result.exit_code == 0, result.output

        wb = openpyxl.load_workbook(output_path, read_only=True)
        assert "Filtered" in wb.sheetnames
        ws_filtered = wb["Filtered"]
        rows = list(ws_filtered.iter_rows(values_only=True))
        # Header + 2 filtered rows
        assert len(rows) == 3
        models_in_sheet = {row[0] for row in rows[1:]}
        assert "NEW" in models_in_sheet
        assert "UK-4216" in models_in_sheet
        wb.close()

    def test_no_filtered_sheet_when_nothing_filtered(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_test_xlsx(
            input_path,
            [("EX4300-48T", "Juniper", "switch")],
        )
        output_path = tmp_path / "output.xlsx"
        runner = CliRunner()

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
        ):
            mock_list.return_value = {"__fallback__": _FakeChecker}
            mock_get.side_effect = (
                lambda name: _FakeChecker if name == "__fallback__" else None
            )
            result = runner.invoke(
                cli,
                [
                    "check",
                    "--input", str(input_path),
                    "--output", str(output_path),
                    "--no-cache",
                ],
            )

        assert result.exit_code == 0, result.output

        wb = openpyxl.load_workbook(output_path, read_only=True)
        assert "Filtered" not in wb.sheetnames
        wb.close()
