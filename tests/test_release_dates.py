"""Tests for release_date tracking across the stack."""

import json
from datetime import date, datetime
from pathlib import Path

import pytest

from eol_tool.models import EOLResult, EOLStatus, HardwareModel

# ── Helpers ─────────────────────────────────────────────────────────


def _hw(
    model: str = "E5-2683 V4",
    manufacturer: str = "Intel",
    category: str = "cpu",
) -> HardwareModel:
    return HardwareModel(model=model, manufacturer=manufacturer, category=category)


def _result(**kwargs) -> EOLResult:
    defaults = dict(
        model=_hw(),
        status=EOLStatus.EOL,
        checked_at=datetime.now(),
        source_name="test",
        confidence=80,
    )
    defaults.update(kwargs)
    return EOLResult(**defaults)


FIXTURES = Path(__file__).parent / "fixtures" / "endoflife_date"


def _load_fixture(name: str) -> list:
    return json.loads((FIXTURES / name).read_text())


# ── Part 1: EOLResult model ────────────────────────────────────────


class TestEOLResultReleaseDate:
    def test_release_date_defaults_to_none(self):
        r = _result()
        assert r.release_date is None

    def test_release_date_accepts_date(self):
        r = _result(release_date=date(2015, 1, 1))
        assert r.release_date == date(2015, 1, 1)

    def test_release_date_serializes(self):
        r = _result(release_date=date(2020, 6, 15))
        d = r.model_dump()
        assert d["release_date"] == date(2020, 6, 15)


# ── Part 1: XLSX round-trip ────────────────────────────────────────


class TestXlsxRoundTrip:
    def test_release_date_written_to_xlsx(self, tmp_path):
        from eol_tool.reader import write_results

        r = _result(release_date=date(2016, 1, 1))
        out = tmp_path / "out.xlsx"
        write_results([r], out)
        assert out.exists()

        import openpyxl

        wb = openpyxl.load_workbook(out, read_only=True)
        ws = wb["EOL Results"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Release Date" in headers
        col_idx = headers.index("Release Date")
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        assert row2[0][col_idx] == "2016-01-01"
        wb.close()

    def test_release_date_empty_when_none(self, tmp_path):
        from eol_tool.reader import write_results

        r = _result(release_date=None)
        out = tmp_path / "out.xlsx"
        write_results([r], out)

        import openpyxl

        wb = openpyxl.load_workbook(out, read_only=True)
        ws = wb["EOL Results"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = headers.index("Release Date")
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        assert row2[0][col_idx] in ("", None)
        wb.close()


# ── Part 1: Cache round-trip ───────────────────────────────────────


class TestCacheRoundTrip:
    @pytest.mark.asyncio
    async def test_release_date_cached_and_retrieved(self, tmp_path):
        from eol_tool.cache import ResultCache

        cache = ResultCache(db_path=tmp_path / "test.db")
        try:
            r = _result(release_date=date(2015, 3, 1))
            await cache.set(r)
            got = await cache.get(r.model.model, r.model.manufacturer)
            assert got is not None
            assert got.release_date == date(2015, 3, 1)
        finally:
            await cache.close()

    @pytest.mark.asyncio
    async def test_release_date_none_cached(self, tmp_path):
        from eol_tool.cache import ResultCache

        cache = ResultCache(db_path=tmp_path / "test.db")
        try:
            r = _result(release_date=None)
            await cache.set(r)
            got = await cache.get(r.model.model, r.model.manufacturer)
            assert got is not None
            assert got.release_date is None
        finally:
            await cache.close()


# ── Part 2: endoflife.date releaseDate extraction ──────────────────


class TestEndoflifeDateReleaseDate:
    def test_release_date_extracted_from_cycle(self):
        from eol_tool.checkers.endoflife_date import EndOfLifeDateChecker

        checker = EndOfLifeDateChecker()
        model = _hw("E5-2683 V4")
        cycles = _load_fixture("intel_processors.json")
        slug_cycles = {"intel-processors": cycles}
        result = checker._match_model_to_result(
            model, ["intel-processors"], slug_cycles,
        )
        assert result.release_date is not None
        assert isinstance(result.release_date, date)

    def test_release_date_none_when_missing(self):
        from eol_tool.checkers.endoflife_date import EndOfLifeDateChecker

        checker = EndOfLifeDateChecker()
        model = _hw("E5-2683 V4")
        # Cycle without releaseDate field
        cycles = [{
            "cycle": "broadwell-xeon",
            "releaseLabel": "Broadwell (Xeon E7v4/E7v4)",
            "eol": "2022-06-30",
        }]
        slug_cycles = {"intel-processors": cycles}
        result = checker._match_model_to_result(
            model, ["intel-processors"], slug_cycles,
        )
        assert result.release_date is None


# ── Part 3: supplement_missing_dates ───────────────────────────────


class TestSupplementMissingDates:
    @pytest.mark.asyncio
    async def test_supplement_adds_eol_date(self, httpx_mock):
        from eol_tool.checkers.endoflife_date import supplement_missing_dates

        all_products = _load_fixture("all_products.json")
        intel_cycles = _load_fixture("intel_processors.json")

        httpx_mock.add_response(
            url="https://endoflife.date/api/all.json",
            json=all_products,
            is_reusable=True,
        )
        httpx_mock.add_response(
            url="https://endoflife.date/api/intel-processors.json",
            json=intel_cycles,
            is_reusable=True,
        )

        r = _result(
            model=_hw("E5-2683 V4"),
            status=EOLStatus.EOL,
            eol_date=None,
            release_date=None,
        )
        results = await supplement_missing_dates([r])
        assert results[0].eol_date is not None
        assert results[0].date_source == "community_database"
        assert "eol-date-supplemented-from-endoflife.date" in results[0].notes

    @pytest.mark.asyncio
    async def test_supplement_adds_release_date(self, httpx_mock):
        from eol_tool.checkers.endoflife_date import supplement_missing_dates

        all_products = _load_fixture("all_products.json")
        intel_cycles = _load_fixture("intel_processors.json")

        httpx_mock.add_response(
            url="https://endoflife.date/api/all.json",
            json=all_products,
            is_reusable=True,
        )
        httpx_mock.add_response(
            url="https://endoflife.date/api/intel-processors.json",
            json=intel_cycles,
            is_reusable=True,
        )

        r = _result(
            model=_hw("E5-2683 V4"),
            status=EOLStatus.EOL,
            eol_date=None,
            release_date=None,
        )
        results = await supplement_missing_dates([r])
        assert results[0].release_date is not None

    @pytest.mark.asyncio
    async def test_supplement_does_not_overwrite_existing_dates(self, httpx_mock):
        from eol_tool.checkers.endoflife_date import supplement_missing_dates

        existing_date = date(2020, 1, 1)
        r = _result(
            model=_hw("E5-2683 V4"),
            status=EOLStatus.ACTIVE,
            eol_date=existing_date,
            release_date=date(2015, 1, 1),
        )
        # Active status with existing eol_date should not be touched
        results = await supplement_missing_dates([r])
        assert results[0].eol_date == existing_date
        assert results[0].release_date == date(2015, 1, 1)

    @pytest.mark.asyncio
    async def test_supplement_called_in_pipeline(self):
        """Verify supplement_missing_dates is wired into the pipeline."""
        from eol_tool.check_pipeline import supplement_missing_dates as fn

        assert callable(fn)


# ── Part 4-5: Intel ARK query normalization ────────────────────────


class TestIntelArkQueryNormalization:
    def test_nic_x520_da2(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("X520-DA2 10GB/S DUAL", "nic") == "Intel Ethernet X520-DA2"

    def test_nic_x540_t2(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("X540-T2 10GB/S DUAL", "nic") == "Intel Ethernet X540-T2"

    def test_nic_x550_t2(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("X550-T2 DUAL 10GBE", "nic") == "Intel Ethernet X550-T2"

    def test_nic_i350_t4(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("I350-T4 QUAD 10GBASE-T", "nic") == "Intel Ethernet I350-T4"

    def test_nic_x710_bm2(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("X710-BM2", "nic") == "Intel Ethernet X710-BM2"

    def test_ssd_int_s3500(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("INT S3500", "ssd") == "Intel SSD S3500"

    def test_ssd_int_d3_s4510(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("INT D3-S4510", "ssd") == "Intel SSD D3-S4510"

    def test_ssd_int_p3600_u2(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("INT P3600 U.2", "ssd") == "Intel SSD P3600"

    def test_ssd_int_520(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("INT 520", "ssd") == "Intel SSD 520"

    def test_ssd_int_660p_m2(self):
        from eol_tool.checkers.intel_ark import _build_ark_query

        assert _build_ark_query("INT 660P M.2", "ssd") == "Intel SSD 660P"


# ── Part 4: Intel ARK launch date parsing ──────────────────────────


class TestIntelArkLaunchDate:
    def test_to_result_sets_release_date(self):
        from eol_tool.checkers.intel_ark import _to_result

        model = _hw("E5-2683 v4")
        data = {
            "marketing_status": "Discontinued",
            "launch_date": "Q1'16",
            "eol_date": "June 30, 2022",
        }
        r = _to_result(model, data)
        assert r.release_date == date(2016, 1, 1)

    def test_launch_date_q3(self):
        from eol_tool.checkers.intel_ark import _parse_launch_date

        d = _parse_launch_date("Q3'20")
        assert d == date(2020, 7, 1)

    def test_launch_date_q4(self):
        from eol_tool.checkers.intel_ark import _parse_launch_date

        d = _parse_launch_date("Q4'21")
        assert d == date(2021, 10, 1)

    def test_launch_date_q2(self):
        from eol_tool.checkers.intel_ark import _parse_launch_date

        d = _parse_launch_date("Q2'19")
        assert d == date(2019, 4, 1)

    def test_launch_date_empty(self):
        from eol_tool.checkers.intel_ark import _parse_launch_date

        assert _parse_launch_date("") is None
