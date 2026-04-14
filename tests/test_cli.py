"""Tests for CLI logging configuration and output flags."""

import logging
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import openpyxl
from click.testing import CliRunner

from eol_tool.cli import cli
from eol_tool.models import EOLResult, EOLStatus


def test_default_log_level_is_warning():
    runner = CliRunner()
    result = runner.invoke(cli, ["list-checkers"])
    assert result.exit_code == 0
    root = logging.getLogger()
    assert root.level == logging.WARNING


def test_log_level_debug():
    runner = CliRunner()
    result = runner.invoke(cli, ["--log-level", "debug", "list-checkers"])
    assert result.exit_code == 0
    root = logging.getLogger()
    assert root.level == logging.DEBUG


def test_verbose_sets_info():
    runner = CliRunner()
    result = runner.invoke(cli, ["-v", "list-checkers"])
    assert result.exit_code == 0
    root = logging.getLogger()
    assert root.level == logging.INFO


def test_log_level_overrides_verbose():
    runner = CliRunner()
    result = runner.invoke(cli, ["-v", "--log-level", "error", "list-checkers"])
    assert result.exit_code == 0
    root = logging.getLogger()
    assert root.level == logging.ERROR


# ── Helpers for output-flag tests ────────────────────────────────────


def _make_test_xlsx(path: Path, rows: list[dict]) -> Path:
    """Create a minimal xlsx for CLI tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Models"
    headers = ["Model", "Manufacturer", "Category", "Condition", "Original_Item"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for ri, row in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=row.get("model", ""))
        ws.cell(row=ri, column=2, value=row.get("manufacturer", ""))
        ws.cell(row=ri, column=3, value=row.get("category", "unknown"))
        ws.cell(row=ri, column=4, value=row.get("condition", ""))
        ws.cell(row=ri, column=5, value=row.get("original_item", ""))
    wb.save(path)
    wb.close()
    return path


def _fake_pipeline_results(models, **kwargs):
    """Return a simple EOLResult for each model."""
    return [
        EOLResult(
            model=m,
            status=EOLStatus.ACTIVE,
            checked_at=datetime.now(),
            source_name="test",
        )
        for m in models
    ]


_PIPELINE_PATH = "eol_tool.check_pipeline.run_check_pipeline"


# ── Output-flag tests ───────────────────────────────────────────────


def test_quiet_flag_suppresses_progress(tmp_path):
    xlsx = _make_test_xlsx(
        tmp_path / "input.xlsx",
        [
            {"model": "X520-DA2", "manufacturer": "Intel", "category": "nic"},
            {"model": "EX4300", "manufacturer": "Juniper", "category": "switch"},
        ],
    )

    async def _fake(models, **kw):
        return _fake_pipeline_results(models)

    runner = CliRunner()
    with patch(_PIPELINE_PATH, side_effect=_fake):
        result = runner.invoke(
            cli, ["check", "--input", str(xlsx), "-q", "--no-cache", "--skip-fallback"],
        )

    assert result.exit_code == 0, result.output
    # -q should suppress "Loaded", "Manufacturers", and "Checking" lines
    assert "Loaded" not in result.output
    assert "Manufacturers" not in result.output
    assert "Checking" not in result.output


def test_show_warnings_lists_individual_rows(tmp_path):
    xlsx = _make_test_xlsx(
        tmp_path / "input.xlsx",
        [
            {"model": "MYSTERY-1", "manufacturer": "", "category": "cpu"},
            {"model": "MYSTERY-2", "manufacturer": "", "category": "cpu"},
            {"model": "X520-DA2", "manufacturer": "Intel", "category": "nic"},
        ],
    )

    runner = CliRunner()
    result = runner.invoke(
        cli, ["check", "--input", str(xlsx), "--show-warnings", "--dry-run"],
    )

    assert result.exit_code == 0, result.output
    # With --show-warnings, each no-manufacturer row should be logged individually
    # Logging goes to stderr via basicConfig
    all_output = result.output
    assert "MYSTERY-1" in all_output
    assert "MYSTERY-2" in all_output


def test_default_batches_warnings(tmp_path):
    xlsx = _make_test_xlsx(
        tmp_path / "input.xlsx",
        [
            {"model": "MYSTERY-1", "manufacturer": "", "category": "cpu"},
            {"model": "MYSTERY-2", "manufacturer": "", "category": "cpu"},
            {"model": "MYSTERY-3", "manufacturer": "", "category": "cpu"},
            {"model": "X520-DA2", "manufacturer": "Intel", "category": "nic"},
        ],
    )

    runner = CliRunner()
    result = runner.invoke(
        cli, ["check", "--input", str(xlsx), "--dry-run"],
    )

    assert result.exit_code == 0, result.output
    all_output = result.output
    # Default: batch summary, not individual model names
    assert "3 rows have no manufacturer" in all_output
    assert "MYSTERY-1" not in all_output
