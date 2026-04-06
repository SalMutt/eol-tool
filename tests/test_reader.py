"""Tests for reader module."""

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

from eol_tool.models import EOLReason, EOLResult, EOLStatus, HardwareModel, RiskCategory
from eol_tool.reader import read_models, write_results

_DATA_FILE = Path(__file__).parent.parent / "data" / "eol_models_cleaned.xlsx"


@pytest.mark.skipif(not _DATA_FILE.exists(), reason="inventory data file not present")
class TestReadModels:
    def test_read_count(self):
        models = read_models(_DATA_FILE)
        assert len(models) > 1000

    def test_no_empty_models(self):
        models = read_models(_DATA_FILE)
        for m in models:
            assert m.model.strip() != ""

    def test_all_have_category(self):
        models = read_models(_DATA_FILE)
        for m in models:
            assert m.category != ""

    def test_spot_check_intel(self):
        models = read_models(_DATA_FILE)
        intel = [m for m in models if m.manufacturer == "Intel"]
        assert len(intel) > 0
        assert any("XEON" in m.model or "E" in m.model for m in intel)

    def test_spot_check_juniper(self):
        models = read_models(_DATA_FILE)
        juniper = [m for m in models if m.manufacturer == "Juniper"]
        assert len(juniper) > 0
        assert all(m.model == m.model.upper() for m in juniper)

    def test_manufacturers_present(self):
        models = read_models(_DATA_FILE)
        mfrs = {m.manufacturer for m in models}
        assert "Intel" in mfrs
        assert "Juniper" in mfrs
        assert "Samsung" in mfrs


class TestWriteResults:
    def test_creates_xlsx(self, tmp_path):
        results = _sample_results()
        out = tmp_path / "output.xlsx"
        write_results(results, out)
        assert out.exists()

    def test_sheet_names(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        assert "EOL Results" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        wb.close()

    def test_header_row(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 17)]
        assert headers[0] == "Model"
        assert headers[4] == "EOL Status"
        assert headers[7] == "Release Date"
        assert headers[8] == "Date Source"
        assert headers[12] == "Original_Item"
        assert headers[13] == "Notes"
        assert headers[14] == "EOL Reason"
        assert headers[15] == "Risk Category"
        wb.close()

    def test_data_rows(self, tmp_path):
        results = _sample_results()
        out = tmp_path / "output.xlsx"
        write_results(results, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        assert ws.cell(row=2, column=1).value == "EX4300-48T"
        assert ws.cell(row=2, column=5).value == "eol"
        wb.close()

    def test_status_coloring(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        eol_cell = ws.cell(row=2, column=5)
        assert eol_cell.fill.fgColor.rgb == "00FFE6E6"
        wb.close()

    def test_frozen_panes(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        assert ws.freeze_panes == "A2"
        wb.close()

    def test_auto_filter(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        assert ws.auto_filter.ref is not None
        wb.close()

    def test_summary_totals(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]
        assert ws.cell(row=3, column=2).value == 3  # total models
        wb.close()

    def test_new_columns_written(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        assert ws.cell(row=2, column=15).value == "community_data"
        assert ws.cell(row=2, column=16).value == "security"
        assert ws.cell(row=3, column=16).value == "support"
        assert ws.cell(row=4, column=16).value == "procurement"
        wb.close()

    def test_risk_category_coloring(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        # security row
        risk_cell = ws.cell(row=2, column=16)
        assert risk_cell.fill.fgColor.rgb == "00FFE0E0"
        # support row
        risk_cell = ws.cell(row=3, column=16)
        assert risk_cell.fill.fgColor.rgb == "00FFF0E0"
        wb.close()

    def test_original_item_column(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results_with_original(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        # Original_Item column should contain the original_item value
        assert ws.cell(row=2, column=13).value == "EX4300-48T-AFI"
        # Row without original_item should be empty
        assert not ws.cell(row=3, column=13).value
        wb.close()

    def test_model_column_uses_original_item(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results_with_original(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["EOL Results"]
        # Model column shows original_item when available
        assert ws.cell(row=2, column=1).value == "EX4300-48T-AFI"
        # Falls back to model when original_item is empty
        assert ws.cell(row=3, column=1).value == "XEON E3-1230V5"
        wb.close()

    def test_summary_risk_counts(self, tmp_path):
        out = tmp_path / "output.xlsx"
        write_results(_sample_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]
        assert ws.cell(row=7, column=1).value == "Risk: Security"
        assert ws.cell(row=7, column=2).value == 1
        assert ws.cell(row=8, column=1).value == "Risk: Support"
        assert ws.cell(row=8, column=2).value == 1
        assert ws.cell(row=9, column=1).value == "Risk: Procurement"
        assert ws.cell(row=9, column=2).value == 1
        wb.close()


def _sample_results() -> list[EOLResult]:
    now = datetime(2025, 6, 1, 12, 0, 0)
    return [
        EOLResult(
            model=HardwareModel(model="EX4300-48T", manufacturer="Juniper", category="switch"),
            status=EOLStatus.EOL,
            eol_date=date(2023, 6, 15),
            checked_at=now,
            confidence=90,
            eol_reason=EOLReason.COMMUNITY_DATA,
            risk_category=RiskCategory.SECURITY,
        ),
        EOLResult(
            model=HardwareModel(model="XEON E3-1230V5", manufacturer="Intel", category="cpu"),
            status=EOLStatus.ACTIVE,
            checked_at=now,
            confidence=80,
            eol_reason=EOLReason.COMMUNITY_DATA,
            risk_category=RiskCategory.SUPPORT,
        ),
        EOLResult(
            model=HardwareModel(model="PM1643A", manufacturer="Samsung", category="ssd"),
            status=EOLStatus.UNKNOWN,
            checked_at=now,
            eol_reason=EOLReason.COMMUNITY_DATA,
            risk_category=RiskCategory.PROCUREMENT,
        ),
    ]


def _sample_results_with_original() -> list[EOLResult]:
    now = datetime(2025, 6, 1, 12, 0, 0)
    return [
        EOLResult(
            model=HardwareModel(
                model="EX4300-48T",
                manufacturer="Juniper",
                category="switch",
                original_item="EX4300-48T-AFI",
            ),
            status=EOLStatus.EOL,
            eol_date=date(2023, 6, 15),
            checked_at=now,
            confidence=90,
            eol_reason=EOLReason.COMMUNITY_DATA,
            risk_category=RiskCategory.SECURITY,
        ),
        EOLResult(
            model=HardwareModel(model="XEON E3-1230V5", manufacturer="Intel", category="cpu"),
            status=EOLStatus.ACTIVE,
            checked_at=now,
            confidence=80,
            eol_reason=EOLReason.COMMUNITY_DATA,
            risk_category=RiskCategory.SUPPORT,
        ),
    ]
