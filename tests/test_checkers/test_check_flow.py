"""Tests for the check command flow."""

from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import pytest
from click.testing import CliRunner

from eol_tool.checker import BaseChecker
from eol_tool.cli import cli
from eol_tool.models import EOLResult, EOLStatus, HardwareModel


class FakeChecker(BaseChecker):
    """A test checker that returns predictable results."""

    manufacturer_name = "testmfr"
    rate_limit = 10

    async def check(self, model: HardwareModel) -> EOLResult:
        return EOLResult(
            model=model,
            status=EOLStatus.EOL,
            checked_at=datetime.now(),
            source_name="fake",
            confidence=90,
            notes="fake-check",
        )


class ErrorChecker(BaseChecker):
    """A checker that raises on every check."""

    manufacturer_name = "errormfr"
    rate_limit = 10

    async def check(self, model: HardwareModel) -> EOLResult:
        raise RuntimeError("intentional test error")


class InitErrorChecker(BaseChecker):
    """A checker that fails on __aenter__."""

    manufacturer_name = "initerrormfr"
    rate_limit = 10

    async def __aenter__(self):
        raise RuntimeError("init failed")

    async def check(self, model: HardwareModel) -> EOLResult:
        raise NotImplementedError


def _write_test_xlsx(path: Path, models: list[tuple[str, str, str]]) -> None:
    """Write a minimal xlsx file with test models."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Models"
    ws.append(["Model", "Manufacturer", "Category", "Condition", "Original_Item"])
    for model, mfr, cat in models:
        ws.append([model, mfr, cat, "", ""])
    wb.save(path)
    wb.close()


@pytest.fixture
def test_xlsx(tmp_path):
    """Create a test xlsx with models from two manufacturers."""
    path = tmp_path / "input.xlsx"
    _write_test_xlsx(
        path,
        [
            ("Model-A1", "Alpha", "cpu"),
            ("Model-A2", "Alpha", "cpu"),
            ("Model-B1", "Beta", "switch"),
        ],
    )
    return path


@pytest.fixture
def runner():
    return CliRunner()


class TestCheckCommandProducesOutput:
    """Test that check command with mocked checkers produces correct output xlsx."""

    def test_check_writes_xlsx(self, runner, test_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
        ):
            mock_list.return_value = {"__fallback__": FakeChecker}
            mock_get.side_effect = lambda name: FakeChecker if name == "__fallback__" else None

            result = runner.invoke(
                cli,
                ["check", "--input", str(test_xlsx), "--output", str(output), "--no-cache"],
            )

        assert result.exit_code == 0, result.output
        assert output.exists()
        assert "Results written to" in result.output

        # Verify the output xlsx has correct data
        import openpyxl

        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb["EOL Results"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 3 data rows
        assert len(rows) == 4
        # Check statuses
        statuses = [row[4] for row in rows[1:]]
        assert all(s == "eol" for s in statuses)
        wb.close()

    def test_check_manufacturer_filter(self, runner, test_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
        ):
            mock_list.return_value = {"__fallback__": FakeChecker}
            mock_get.side_effect = lambda name: FakeChecker if name == "__fallback__" else None

            result = runner.invoke(
                cli,
                [
                    "check",
                    "--input", str(test_xlsx),
                    "--output", str(output),
                    "--manufacturer", "Alpha",
                    "--no-cache",
                ],
            )

        assert result.exit_code == 0, result.output

        import openpyxl

        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb["EOL Results"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 2 Alpha models only
        assert len(rows) == 3
        manufacturers = [row[1] for row in rows[1:]]
        assert all(m == "Alpha" for m in manufacturers)
        wb.close()


class TestCacheIntegration:
    """Test that cache integration works."""

    def test_second_run_uses_cache(self, runner, test_xlsx, tmp_path):
        output1 = tmp_path / "output1.xlsx"
        output2 = tmp_path / "output2.xlsx"
        db_path = tmp_path / "cache.db"

        call_count = 0
        original_check = FakeChecker.check

        class CountingChecker(FakeChecker):
            async def check(self, model):
                nonlocal call_count
                call_count += 1
                return await original_check(self, model)

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
            patch("eol_tool.cache._DEFAULT_DB", db_path),
        ):
            mock_list.return_value = {"__fallback__": CountingChecker}
            mock_get.side_effect = (
                lambda name: CountingChecker if name == "__fallback__" else None
            )

            # First run: all models checked
            result1 = runner.invoke(
                cli,
                ["check", "--input", str(test_xlsx), "--output", str(output1)],
            )
            assert result1.exit_code == 0, result1.output
            assert call_count == 3

            # Second run: should use cache
            call_count = 0
            result2 = runner.invoke(
                cli,
                ["check", "--input", str(test_xlsx), "--output", str(output2)],
            )
            assert result2.exit_code == 0, result2.output
            assert call_count == 0, f"Expected 0 checks on second run, got {call_count}"
            assert "cached" in result2.output

    def test_no_cache_bypasses_cache(self, runner, test_xlsx, tmp_path):
        output1 = tmp_path / "output1.xlsx"
        output2 = tmp_path / "output2.xlsx"
        db_path = tmp_path / "cache.db"

        call_count = 0
        original_check = FakeChecker.check

        class CountingChecker(FakeChecker):
            async def check(self, model):
                nonlocal call_count
                call_count += 1
                return await original_check(self, model)

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
            patch("eol_tool.cache._DEFAULT_DB", db_path),
        ):
            mock_list.return_value = {"__fallback__": CountingChecker}
            mock_get.side_effect = (
                lambda name: CountingChecker if name == "__fallback__" else None
            )

            # First run
            result1 = runner.invoke(
                cli,
                ["check", "--input", str(test_xlsx), "--output", str(output1)],
            )
            assert result1.exit_code == 0, result1.output

            # Second run with --no-cache: should check all again
            call_count = 0
            result2 = runner.invoke(
                cli,
                [
                    "check",
                    "--input", str(test_xlsx),
                    "--output", str(output2),
                    "--no-cache",
                ],
            )
            assert result2.exit_code == 0, result2.output
            assert call_count == 3, f"Expected 3 checks with --no-cache, got {call_count}"


class TestExceptionHandling:
    """Test that checker exceptions are caught and result in UNKNOWN status."""

    def test_per_model_exception_returns_unknown(self, runner, test_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
        ):
            mock_list.return_value = {"__fallback__": ErrorChecker}
            mock_get.side_effect = lambda name: ErrorChecker if name == "__fallback__" else None

            result = runner.invoke(
                cli,
                ["check", "--input", str(test_xlsx), "--output", str(output), "--no-cache"],
            )

        assert result.exit_code == 0, result.output
        assert output.exists()

        import openpyxl

        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb["EOL Results"]
        rows = list(ws.iter_rows(values_only=True))
        statuses = [row[4] for row in rows[1:]]
        assert all(s == "unknown" for s in statuses)
        wb.close()

    def test_checker_init_failure_returns_unknown(self, runner, test_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
        ):
            mock_list.return_value = {"__fallback__": InitErrorChecker}
            mock_get.side_effect = (
                lambda name: InitErrorChecker if name == "__fallback__" else None
            )

            result = runner.invoke(
                cli,
                ["check", "--input", str(test_xlsx), "--output", str(output), "--no-cache"],
            )

        assert result.exit_code == 0, result.output
        assert output.exists()

        import openpyxl

        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb["EOL Results"]
        rows = list(ws.iter_rows(values_only=True))
        statuses = [row[4] for row in rows[1:]]
        assert all(s == "unknown" for s in statuses)
        wb.close()


class TestSummaryOutput:
    """Test the summary output."""

    def test_summary_table_printed(self, runner, test_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
        ):
            mock_list.return_value = {"__fallback__": FakeChecker}
            mock_get.side_effect = lambda name: FakeChecker if name == "__fallback__" else None

            result = runner.invoke(
                cli,
                ["check", "--input", str(test_xlsx), "--output", str(output), "--no-cache"],
            )

        assert result.exit_code == 0, result.output
        assert "Manufacturer" in result.output
        assert "Alpha" in result.output
        assert "Beta" in result.output
        assert "Total" in result.output
        assert "Security" in result.output

    def test_progress_messages(self, runner, test_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"

        with (
            patch("eol_tool.check_pipeline.get_checker") as mock_get,
            patch("eol_tool.cli._list_checkers") as mock_list,
        ):
            mock_list.return_value = {"__fallback__": FakeChecker}
            mock_get.side_effect = lambda name: FakeChecker if name == "__fallback__" else None

            result = runner.invoke(
                cli,
                ["check", "--input", str(test_xlsx), "--output", str(output), "--no-cache"],
            )

        assert result.exit_code == 0, result.output
        assert "Checking Alpha: 2 models... done" in result.output
        assert "Checking Beta: 1 models... done" in result.output
