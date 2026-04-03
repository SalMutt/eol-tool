"""Tests for --retry-unknowns flag on the check command."""

from datetime import date, datetime
from unittest.mock import patch

import openpyxl
import pytest
from click.testing import CliRunner

from eol_tool.checker import BaseChecker
from eol_tool.cli import cli
from eol_tool.models import EOLReason, EOLResult, EOLStatus, HardwareModel, RiskCategory
from eol_tool.reader import read_results, split_results_for_retry, write_results


def _make_previous_results() -> list[EOLResult]:
    """Build 10 results: 5 EOL, 2 ACTIVE, 2 UNKNOWN, 1 NOT_FOUND."""
    now = datetime(2025, 7, 1, 12, 0, 0)
    return [
        # --- 5 EOL models (should stay as-is) ---
        EOLResult(
            model=HardwareModel(model="EX4300-48T", manufacturer="Juniper", category="switch"),
            status=EOLStatus.EOL,
            eol_date=date(2023, 6, 15),
            checked_at=now, confidence=95, source_name="juniper",
            eol_reason=EOLReason.MANUFACTURER_DECLARED, risk_category=RiskCategory.SECURITY,
            date_source="manufacturer_confirmed",
        ),
        EOLResult(
            model=HardwareModel(model="WS-C3750X-48T", manufacturer="Cisco", category="switch"),
            status=EOLStatus.EOL,
            eol_date=date(2020, 7, 31),
            checked_at=now, confidence=95, source_name="cisco",
            eol_reason=EOLReason.MANUFACTURER_DECLARED, risk_category=RiskCategory.SECURITY,
            date_source="manufacturer_confirmed",
        ),
        EOLResult(
            model=HardwareModel(model="XEON E5-2680V4", manufacturer="Intel", category="cpu"),
            status=EOLStatus.EOL,
            eol_date=date(2022, 10, 1),
            checked_at=now, confidence=95, source_name="intel",
            eol_reason=EOLReason.MANUFACTURER_DECLARED, risk_category=RiskCategory.INFORMATIONAL,
        ),
        EOLResult(
            model=HardwareModel(model="PM1643A", manufacturer="Samsung", category="ssd"),
            status=EOLStatus.EOL,
            checked_at=now, confidence=85, source_name="endoflife.date",
            eol_reason=EOLReason.COMMUNITY_DATA, risk_category=RiskCategory.SUPPORT,
        ),
        EOLResult(
            model=HardwareModel(model="N9K-C93180YC-EX", manufacturer="Cisco", category="switch"),
            status=EOLStatus.EOL_ANNOUNCED,
            eol_date=date(2026, 1, 1),
            checked_at=now, confidence=90, source_name="cisco",
            eol_reason=EOLReason.MANUFACTURER_DECLARED, risk_category=RiskCategory.PROCUREMENT,
        ),
        # --- 2 ACTIVE models (should stay as-is) ---
        EOLResult(
            model=HardwareModel(model="MX204", manufacturer="Juniper", category="router"),
            status=EOLStatus.ACTIVE,
            checked_at=now, confidence=85, source_name="juniper",
        ),
        EOLResult(
            model=HardwareModel(model="ST4000NM0035", manufacturer="Seagate", category="hdd"),
            status=EOLStatus.ACTIVE,
            checked_at=now, confidence=70, source_name="seagate",
        ),
        # --- 2 UNKNOWN models (should be re-checked) ---
        EOLResult(
            model=HardwareModel(model="PM883", manufacturer="Samsung", category="ssd"),
            status=EOLStatus.UNKNOWN,
            checked_at=now, confidence=50, source_name="samsung",
        ),
        EOLResult(
            model=HardwareModel(model="KSM64R52BD4", manufacturer="Kingston", category="memory"),
            status=EOLStatus.UNKNOWN,
            checked_at=now, confidence=30, source_name="kingston",
        ),
        # --- 1 NOT_FOUND model (should be re-checked) ---
        EOLResult(
            model=HardwareModel(model="X11SPL-F", manufacturer="Supermicro", category="mainboard"),
            status=EOLStatus.NOT_FOUND,
            checked_at=now, confidence=0, source_name="",
        ),
    ]


@pytest.fixture(scope="session")
def previous_results_xlsx(tmp_path_factory):
    """Write the 10-model previous results fixture to an xlsx."""
    tmpdir = tmp_path_factory.mktemp("retry_fixtures")
    path = tmpdir / "previous_results.xlsx"
    write_results(_make_previous_results(), path)
    return path


@pytest.fixture
def runner():
    return CliRunner()


class FakeResolveChecker(BaseChecker):
    """Returns EOL for every model — simulates all unknowns getting resolved."""

    manufacturer_name = "testmfr"
    rate_limit = 10

    async def check(self, model: HardwareModel) -> EOLResult:
        return EOLResult(
            model=model,
            status=EOLStatus.EOL,
            eol_date=date(2025, 1, 1),
            checked_at=datetime.now(),
            source_name="fake-resolve",
            confidence=90,
            eol_reason=EOLReason.TECHNOLOGY_GENERATION,
            risk_category=RiskCategory.SUPPORT,
        )


class FakeStillUnknownChecker(BaseChecker):
    """Returns UNKNOWN for every model — simulates no progress."""

    manufacturer_name = "testmfr"
    rate_limit = 10

    async def check(self, model: HardwareModel) -> EOLResult:
        return EOLResult(
            model=model,
            status=EOLStatus.UNKNOWN,
            checked_at=datetime.now(),
            source_name="fake-still-unknown",
            confidence=10,
        )


def _mock_pipeline(checker_cls):
    """Context managers that wire a FakeChecker into the check pipeline."""
    return (
        patch("eol_tool.check_pipeline.get_checker",
              side_effect=lambda name: checker_cls if name == "__fallback__" else None),
        patch("eol_tool.check_pipeline.get_checkers", return_value=[]),
        patch("eol_tool.cli._list_checkers",
              return_value={"__fallback__": checker_cls}),
    )


class TestSplitResultsForRetry:
    """Tests for the reader helper that splits results."""

    def test_split_counts(self, previous_results_xlsx):
        classified, retry = split_results_for_retry(previous_results_xlsx)
        assert len(classified) == 7  # 5 EOL + 2 ACTIVE
        assert len(retry) == 3  # 2 UNKNOWN + 1 NOT_FOUND

    def test_classified_statuses(self, previous_results_xlsx):
        classified, _ = split_results_for_retry(previous_results_xlsx)
        for r in classified:
            assert r.status in (EOLStatus.EOL, EOLStatus.EOL_ANNOUNCED, EOLStatus.ACTIVE)

    def test_retry_are_hardware_models(self, previous_results_xlsx):
        _, retry = split_results_for_retry(previous_results_xlsx)
        for m in retry:
            assert isinstance(m, HardwareModel)

    def test_manufacturer_filter(self, previous_results_xlsx):
        classified, retry = split_results_for_retry(previous_results_xlsx, manufacturer="Samsung")
        # Only the Samsung UNKNOWN should be retried
        assert len(retry) == 1
        assert retry[0].manufacturer == "Samsung"
        # The other unknowns (Kingston, Supermicro) should be in classified
        assert len(classified) == 9


class TestRetryUnknownsOnlyPassesUnknownsToCheckers:
    """Only UNKNOWN and NOT_FOUND models should be sent to checkers."""

    def test_only_unknowns_checked(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        checked_models = []

        class TrackingChecker(BaseChecker):
            manufacturer_name = "testmfr"
            rate_limit = 10

            async def check(self, model: HardwareModel) -> EOLResult:
                checked_models.append(model.model)
                return EOLResult(
                    model=model, status=EOLStatus.ACTIVE,
                    checked_at=datetime.now(), source_name="tracking", confidence=80,
                )

        with (
            patch("eol_tool.check_pipeline.get_checker",
                  side_effect=lambda name: TrackingChecker if name == "__fallback__" else None),
            patch("eol_tool.check_pipeline.get_checkers", return_value=[]),
            patch("eol_tool.cli._list_checkers",
                  return_value={"__fallback__": TrackingChecker}),
        ):
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])

        assert result.exit_code == 0, result.output
        assert sorted(checked_models) == sorted(["PM883", "KSM64R52BD4", "X11SPL-F"])


class TestRetryUnknownsOutputIntegrity:
    """Tests that the merged output is complete and correct."""

    def test_total_model_count_preserved(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        assert result.exit_code == 0, result.output
        output_results = read_results(output)
        assert len(output_results) == 10

    def test_classified_models_unchanged(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        output_results = read_results(output)
        by_model = {r.model.model: r for r in output_results}
        # Spot-check preserved fields on a classified model
        ex4300 = by_model["EX4300-48T"]
        assert ex4300.status == EOLStatus.EOL
        assert ex4300.eol_date == date(2023, 6, 15)
        assert ex4300.confidence == 95
        assert ex4300.source_name == "juniper"
        assert ex4300.risk_category == RiskCategory.SECURITY

    def test_all_columns_preserved_for_classified(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb["EOL Results"]
        headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        assert "EOL Reason" in headers
        assert "Risk Category" in headers
        assert "Date Source" in headers
        # Find the Cisco EOL row and verify all columns present
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "WS-C3750X-48T":
                assert row[4] == "eol"  # status
                assert row[14] == "security"  # risk category
                break
        wb.close()

    def test_all_unknowns_resolved(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        assert result.exit_code == 0, result.output
        output_results = read_results(output)
        unknowns = [r for r in output_results
                     if r.status in (EOLStatus.UNKNOWN, EOLStatus.NOT_FOUND)]
        assert len(unknowns) == 0

    def test_some_unknowns_remain(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeStillUnknownChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        assert result.exit_code == 0, result.output
        output_results = read_results(output)
        unknowns = [r for r in output_results
                     if r.status in (EOLStatus.UNKNOWN, EOLStatus.NOT_FOUND)]
        assert len(unknowns) == 3


class TestRetryUnknownsSummaryMessages:
    """Tests for the CLI summary output."""

    def test_retry_summary_message(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        assert "Retrying 3 unknown/not-found models" in result.output
        assert "skipping 7 already classified" in result.output

    def test_resolved_count_message(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        assert "Resolved 3 of 3 unknowns" in result.output
        assert "0 remain" in result.output

    def test_partial_resolved_message(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeStillUnknownChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        assert "Resolved 0 of 3 unknowns" in result.output
        assert "3 remain" in result.output


class TestRetryUnknownsFlags:
    """Tests for flag interactions."""

    def test_no_input_required(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        # Should succeed without --input
        assert result.exit_code == 0, result.output

    def test_input_required_without_retry(self, runner):
        result = runner.invoke(cli, ["check", "--output", "/tmp/out.xlsx"])
        assert result.exit_code != 0
        assert "--input is required" in result.output

    def test_manufacturer_filter_with_retry(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        checked_models = []

        class TrackingChecker(BaseChecker):
            manufacturer_name = "testmfr"
            rate_limit = 10

            async def check(self, model: HardwareModel) -> EOLResult:
                checked_models.append(model.model)
                return EOLResult(
                    model=model, status=EOLStatus.ACTIVE,
                    checked_at=datetime.now(), source_name="tracking", confidence=80,
                )

        with (
            patch("eol_tool.check_pipeline.get_checker",
                  side_effect=lambda name: TrackingChecker if name == "__fallback__" else None),
            patch("eol_tool.check_pipeline.get_checkers", return_value=[]),
            patch("eol_tool.cli._list_checkers",
                  return_value={"__fallback__": TrackingChecker}),
        ):
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
                "--manufacturer", "Samsung",
            ])

        assert result.exit_code == 0, result.output
        # Only Samsung's UNKNOWN (PM883) should have been checked
        assert checked_models == ["PM883"]
        # Output should still have all 10 models
        output_results = read_results(output)
        assert len(output_results) == 10

    def test_diff_with_retry(self, runner, previous_results_xlsx, tmp_path):
        output = tmp_path / "output.xlsx"
        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
            ])
        assert result.exit_code == 0, result.output
        # Diff is auto-generated against the retry file; should show changes
        assert "changes detected" in result.output

    def test_diff_flag_overrides_auto_diff(self, runner, previous_results_xlsx, tmp_path):
        """--diff with a specific path should diff against that file, not the retry file."""
        output = tmp_path / "output.xlsx"
        # First create a separate baseline
        baseline = tmp_path / "baseline.xlsx"
        write_results(_make_previous_results(), baseline)

        p1, p2, p3 = _mock_pipeline(FakeResolveChecker)
        with p1, p2, p3:
            result = runner.invoke(cli, [
                "check", "--retry-unknowns", str(previous_results_xlsx),
                "--output", str(output), "--no-cache",
                "--diff", str(baseline),
            ])
        assert result.exit_code == 0, result.output
        assert "changes detected" in result.output
