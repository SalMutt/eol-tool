"""Read hardware models from spreadsheets and write results."""

import logging
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import EOLResult, HardwareModel
from .normalizer import normalize_model

logger = logging.getLogger(__name__)

_EXPECTED_HEADERS = {"model", "manufacturer", "category", "condition", "original_item"}

_RESULT_COLUMNS = [
    "Model",
    "Manufacturer",
    "Category",
    "Condition",
    "EOL Status",
    "EOL Date",
    "EOS Date",
    "Release Date",
    "Date Source",
    "Confidence",
    "Source",
    "Checked At",
    "Original_Item",
    "Notes",
    "EOL Reason",
    "Risk Category",
]

_STATUS_STYLES = {
    "eol": (Font(color="FF0000"), PatternFill(fgColor="FFE6E6", fill_type="solid")),
    "eol_announced": (Font(color="FF8C00"), PatternFill(fgColor="FFF3E0", fill_type="solid")),
    "active": (Font(color="008000"), PatternFill(fgColor="E6FFE6", fill_type="solid")),
    "unknown": (Font(color="666666"), PatternFill(fgColor="F0F0F0", fill_type="solid")),
    "not_found": (Font(color="999999"), PatternFill(fgColor="F5F5F5", fill_type="solid")),
}

_RISK_STYLES = {
    "security": (Font(color="8B0000"), PatternFill(fgColor="FFE0E0", fill_type="solid")),
    "support": (Font(color="FF8C00"), PatternFill(fgColor="FFF0E0", fill_type="solid")),
    "procurement": (Font(color="0000CD"), PatternFill(fgColor="E0E8FF", fill_type="solid")),
    "informational": (Font(color="808080"), PatternFill(fgColor="F0F0F0", fill_type="solid")),
}

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fgColor="2F5496", fill_type="solid")

_DEFAULT_WIDTHS = {
    "Model": 30,
    "Manufacturer": 18,
    "Category": 14,
    "Condition": 12,
    "EOL Status": 16,
    "EOL Date": 14,
    "EOS Date": 14,
    "Release Date": 14,
    "Date Source": 22,
    "Confidence": 12,
    "Source": 30,
    "Checked At": 22,
    "Original_Item": 30,
    "Notes": 40,
    "EOL Reason": 22,
    "Risk Category": 16,
}

_DATE_SOURCE_LABELS = {
    "manufacturer_confirmed": "Manufacturer Confirmed",
    "community_database": "Community Database",
    "none": "Not Available",
}

# Generation-estimate sources from generation_dates.csv map to
# "Generation Estimate" in the output.  Use .get() fallback for
# any source value not in _DATE_SOURCE_LABELS.
_GENERATION_ESTIMATE_SOURCES = frozenset({
    "intel-ark-historical", "amd-historical", "jedec-standard",
    "sata-io-standard", "nvme-standard", "t10-standard",
    "nvidia-historical", "mellanox-historical", "intel-historical",
    "juniper-eol-notices", "supermicro-historical", "dell-support",
    "broadcom-historical", "broadcom-acquisition", "toshiba-acquisition",
    "evga-announcement", "wd-acquisition", "seagate-historical",
    "kingston-historical", "solidigm-historical", "kioxia-historical",
    "wd-historical", "asrock-historical", "asus-historical",
    "cisco-eol-notices", "arista-historical", "transcend-historical",
    "mushkin-historical", "samsung-historical", "sk-hynix-historical",
    "micron-historical", "hpe-historical", "generation-estimate",
    "juniper-support-contract", "juniper-psu-product",
    "intel-nic-product", "corsair-ddr5", "asus-server-gen",
    "asus-board-product", "asrock-board-product", "adaptec-raid-gen",
    "supermicro-board-product", "supermicro-nic-product",
    "supermicro-riser-product", "transcend-memory-product",
    "transcend-ssd-product", "ibm-raid-product", "samsung-ssd-product",
    "mushkin-ssd-product", "pny-memory-product", "pny-ssd-product",
    "msi-ssd-product", "adata-ssd-product", "arista-switch-gen",
    "mellanox-nic-product", "wd-product-line", "micron-ssd-product",
    "nvidia-gpu-product", "gigabyte-ssd-product", "axiom-generic",
    "atech-generic", "kingston-ssd-product", "hp-enterprise-drive",
    "chenbro-chassis", "dynatron-heatsink", "generic-psu",
    "broadcom-raid-gen", "broadcom-nic-product", "dell-raid-gen",
    "dell-optic", "dell-board", "solidigm-ssd-product",
    "crucial-ssd-product", "gigabyte-board-product",
    "seagate-model-era", "seagate-product-line", "toshiba-product-gen",
    "samsung-memory-gen", "commodity-optic",
    "intel-scalable-1st-gen", "intel-ssd-historical",
    "juniper-mx-component", "juniper-qfx-component",
    "passive-infrastructure",
})


def _date_source_label(source: str) -> str:
    """Convert a date_source value to a human-readable label."""
    label = _DATE_SOURCE_LABELS.get(source)
    if label:
        return label
    if source in _GENERATION_ESTIMATE_SOURCES:
        return "Generation Estimate"
    return "Not Available"


# ── Manufacturer auto-detection rules ────────────────────────────────
# Each rule is (compiled_regex, manufacturer_name).
# Rules are tested case-insensitively against the model string.

_MANUFACTURER_RULES: list[tuple[re.Pattern[str], str]] = [
    # Juniper networking
    (re.compile(r"^(?:EX|QFX|SRX|MX)[-\d]", re.IGNORECASE), "Juniper"),
    (re.compile(
        r"^(?:CHAS-|JPSU-|FFANTRAY-|RE-|MPC|MIC[-\d]|PF-|JNP-|PWR-MX|SCBE|QFXC|SP-FXP)",
        re.IGNORECASE,
    ), "Juniper"),
    # Supermicro
    (re.compile(
        r"^(?:CSE-|X(?:9|1[0-3])|H1[2-4]|AOC-|RSC-|BPN-|SNK-|PIO-|MBD|SYS-)",
        re.IGNORECASE,
    ), "Supermicro"),
    # Seagate (ST prefix + digits)
    (re.compile(r"^ST\d{3,}[A-Z]", re.IGNORECASE), "Seagate"),
    # HGST (Hitachi)
    (re.compile(r"\bHGST\b", re.IGNORECASE), "Hitachi"),
    # WD (including WD/ slash variant)
    (re.compile(r"^(?:WD[\d/ ]|WUS|WUSTR)", re.IGNORECASE), "WD"),
    # Kingston memory
    (re.compile(r"^(?:KTD-|KVR|KSM|KTL-|K\dA)", re.IGNORECASE), "Kingston"),
    # Micron (including MT memory part numbers and MTFDDAK/MTFDKC SSD ordering codes)
    (re.compile(r"^(?:MTA|MEM-DR|MT\d|MTFD)", re.IGNORECASE), "Micron"),
    # Samsung (including MZ7xx SATA SSDs and M3xx server memory)
    (re.compile(r"^(?:MZ[A-Z0-9-]|PM\d|SM\d|M3[289]\d)", re.IGNORECASE), "Samsung"),
    # Intel SSDs
    (re.compile(r"^(?:SSDPE|SSDSC|DC\s*P\d|D3-S\d)", re.IGNORECASE), "Intel"),
    # Intel NICs
    (re.compile(r"^(?:X520-|X540-|X550-|X710-|X722-|XXV710-|I350-|E810-)", re.IGNORECASE), "Intel"),
    # Intel boxed/tray processors (BX80xxx, CD80xxx ordering codes)
    (re.compile(r"^(?:BX80|CD80)\d", re.IGNORECASE), "Intel"),
    # Cisco (wireless, ASA, WS-/C-series)
    (re.compile(r"^AIR-", re.IGNORECASE), "Cisco"),
    (re.compile(r"^ASA\d", re.IGNORECASE), "Cisco"),
    # Dell
    (re.compile(r"^POWEREDGE", re.IGNORECASE), "Dell"),
    (re.compile(r"^R\d{3}[a-z]*$", re.IGNORECASE), "Dell"),
    # Crucial memory/SSDs (CT part numbers)
    (re.compile(r"^CT\d", re.IGNORECASE), "Crucial"),
    # Broadcom (including MegaRAID numeric part numbers)
    (re.compile(r"^(?:BCM|N2)", re.IGNORECASE), "Broadcom"),
    (re.compile(r"^\d{4}-\d+I\b", re.IGNORECASE), "Broadcom"),
    # AMD
    (re.compile(r"^(?:EPYC|RYZEN|RADEON)\b", re.IGNORECASE), "AMD"),
    (re.compile(r"^MI\d{2,3}", re.IGNORECASE), "AMD"),
    # NVIDIA (including VCQ Quadro part numbers)
    (re.compile(r"^P\d{4}$", re.IGNORECASE), "NVIDIA"),
    (re.compile(r"^RTX", re.IGNORECASE), "NVIDIA"),
    (re.compile(r"^T\d{3,4}$", re.IGNORECASE), "NVIDIA"),
    (re.compile(r"^A\d{3,4}$", re.IGNORECASE), "NVIDIA"),
    (re.compile(r"^VCQ", re.IGNORECASE), "NVIDIA"),
    # Zotac GPUs (ZT- part numbers)
    (re.compile(r"^ZT-[A-Z]", re.IGNORECASE), "Zotac"),
    # Solidigm SSDs (D5-P series, formerly Intel)
    (re.compile(r"^D5-P\d", re.IGNORECASE), "Solidigm"),
    # Intel Xeon E-series (E3-1230, E5-2699, E-2136, etc.)
    (re.compile(r"^E[3-7]-\d{4}", re.IGNORECASE), "Intel"),
    (re.compile(r"^E-\d{4}", re.IGNORECASE), "Intel"),
    # Intel Xeon number-first format (4110 SILVER XEON, 6146 GOLD)
    (re.compile(
        r"^\d{4}\w?\s+(?:SILVER|GOLD|BRONZE|PLATINUM)", re.IGNORECASE,
    ), "Intel"),
    # Brocade switches
    (re.compile(r"^(?:ACS|FLS|FCX|FI-)\w", re.IGNORECASE), "Brocade"),
    # Transcend (TS part numbers)
    (re.compile(r"^TS\d+[A-Z]", re.IGNORECASE), "Transcend"),
    # SK Hynix memory (HMA, HMT, HMAA, etc.)
    (re.compile(r"^HM[A-Z]", re.IGNORECASE), "SK Hynix"),
    # NVIDIA GPU products (substring match before brand names)
    (re.compile(r"\bNVIDIA\b", re.IGNORECASE), "NVIDIA"),
    (re.compile(r"\bTESLA\s+[KMPVAT]\d", re.IGNORECASE), "NVIDIA"),
    (re.compile(r"\bGEFORCE\b", re.IGNORECASE), "NVIDIA"),
    (re.compile(r"\bQUADRO\b", re.IGNORECASE), "NVIDIA"),
    # Dynatron heatsinks
    (re.compile(r"\bDYNATRON\b", re.IGNORECASE), "Dynatron"),
    # HP/HPE (match "HP " or "HPE " before a model number to avoid false positives)
    (re.compile(r"\bHPE\b", re.IGNORECASE), "HPE"),
    (re.compile(r"\bHP\s+[A-Z]{2}\d", re.IGNORECASE), "HPE"),
    # Chenbro
    (re.compile(r"\bCHENBRO\b", re.IGNORECASE), "Chenbro"),
    # ── Brand name substring matches ──────────────────────────────────
    (re.compile(r"\bSUPERMICRO\b", re.IGNORECASE), "Supermicro"),
    (re.compile(r"\bCSE-", re.IGNORECASE), "Supermicro"),
    (re.compile(r"\bCISCO\b", re.IGNORECASE), "Cisco"),
    (re.compile(r"\bZOTAC\b", re.IGNORECASE), "Zotac"),
    (re.compile(r"\bSEAGATE\b", re.IGNORECASE), "Seagate"),
    (re.compile(r"\bSAMSUNG\b", re.IGNORECASE), "Samsung"),
    (re.compile(r"\bKINGSTON\b", re.IGNORECASE), "Kingston"),
    (re.compile(r"\bBROADCOM\b", re.IGNORECASE), "Broadcom"),
    (re.compile(r"\bBROCADE\b", re.IGNORECASE), "Brocade"),
    (re.compile(r"\bMELLANOX\b", re.IGNORECASE), "Mellanox"),
    (re.compile(r"\bCORSAIR\b", re.IGNORECASE), "Corsair"),
    (re.compile(r"\b(?:SK\s+)?HYNIX\b", re.IGNORECASE), "SK Hynix"),
    (re.compile(r"\bSANDISK\b", re.IGNORECASE), "SanDisk"),
    (re.compile(r"\bASROCK\b", re.IGNORECASE), "ASRock"),
    (re.compile(r"\bTOSHIBA\b", re.IGNORECASE), "Toshiba"),
    (re.compile(r"\bKIOXIA\b", re.IGNORECASE), "Kioxia"),
    (re.compile(r"\bGIGABYTE\b", re.IGNORECASE), "Gigabyte"),
    (re.compile(r"\bCRUCIAL\b", re.IGNORECASE), "Crucial"),
    (re.compile(r"\bSOLIDIGM\b", re.IGNORECASE), "Solidigm"),
    (re.compile(r"\bDELL\b", re.IGNORECASE), "Dell"),
    (re.compile(r"\bIBM\b", re.IGNORECASE), "IBM"),
    (re.compile(r"\bASUS\b", re.IGNORECASE), "ASUS"),
    (re.compile(r"\bJUNIPER\b", re.IGNORECASE), "Juniper"),
    (re.compile(r"\bHITACHI\b", re.IGNORECASE), "Hitachi"),
    (re.compile(r"\bPNY\b", re.IGNORECASE), "PNY"),
    (re.compile(r"\bEVGA\b", re.IGNORECASE), "EVGA"),
    (re.compile(r"\bOCZ\b", re.IGNORECASE), "OCZ"),
    (re.compile(r"\bARISTA\b", re.IGNORECASE), "Arista"),
    (re.compile(r"\bADATA\b", re.IGNORECASE), "ADATA"),
    (re.compile(r"\bADAPTEC\b", re.IGNORECASE), "Adaptec"),
    (re.compile(r"\bTURBOIRON\b", re.IGNORECASE), "Brocade"),
    (re.compile(r"\bSERVERIRON\b", re.IGNORECASE), "Brocade"),
    # ── Brand prefix matches ──────────────────────────────────────────
    (re.compile(r"^AMD\b", re.IGNORECASE), "AMD"),
    (re.compile(r"^INTEL\b", re.IGNORECASE), "Intel"),
    # ── Abbreviated brand prefixes ────────────────────────────────────
    (re.compile(r"^INT\b(?!ERN)", re.IGNORECASE), "Intel"),
    (re.compile(r"^TOS\b", re.IGNORECASE), "Toshiba"),
    (re.compile(r"^SAM\b", re.IGNORECASE), "Samsung"),
    (re.compile(r"^KNG\b", re.IGNORECASE), "Kingston"),
    (re.compile(r"^LSI\b", re.IGNORECASE), "Broadcom"),
    (re.compile(r"^MIC\b", re.IGNORECASE), "Micron"),
    (re.compile(r"^CRU\b", re.IGNORECASE), "Crucial"),
    (re.compile(r"^CP\d", re.IGNORECASE), "Corsair"),
    (re.compile(r"^MUS\b", re.IGNORECASE), "Mushkin"),
    (re.compile(r"^TRAN\b", re.IGNORECASE), "Transcend"),
    (re.compile(r"^ASU\b", re.IGNORECASE), "ASUS"),
    (re.compile(r"^RS\d{3}", re.IGNORECASE), "ASUS"),
    (re.compile(r"^SM\s", re.IGNORECASE), "Supermicro"),
    (re.compile(r"^SMC\b", re.IGNORECASE), "Supermicro"),
    (re.compile(r"^MSI\b", re.IGNORECASE), "MSI"),
    (re.compile(r"^AXIOM\b", re.IGNORECASE), "Axiom"),
    (re.compile(r"^A-TECH\b", re.IGNORECASE), "A-Tech"),
    (re.compile(r"^ASR\b", re.IGNORECASE), "ASRock"),
    (re.compile(r"^SPC\d", re.IGNORECASE), "ASRock"),
    (re.compile(r"^MNT[\s-]", re.IGNORECASE), "Juniper"),
    (re.compile(r"^FS-", re.IGNORECASE), "FS.com"),
    (re.compile(r"\bEDGEVANA\b", re.IGNORECASE), "Edgevana"),
]

# Category-specific rules (checked first for higher specificity)
_CATEGORY_MANUFACTURER_RULES: list[tuple[str, re.Pattern[str], str]] = [
    ("switch", re.compile(r"^(?:EX|QFX)", re.IGNORECASE), "Juniper"),
    ("memory", re.compile(r"^MTA", re.IGNORECASE), "Micron"),
    ("memory", re.compile(r"^(?:KTD-|KVR|KSM)", re.IGNORECASE), "Kingston"),
    # FS.com optics use numeric catalog IDs as MPNs
    ("optic", re.compile(r"^\d+$"), "FS.com"),
]


# Regexes for stripping category/condition/capacity prefixes in detection
_DETECTION_PREFIX_RE = re.compile(
    r"^[A-Z][A-Z /]+:(NEW|USED|REFURBISHED):", re.IGNORECASE,
)
_DETECTION_CAPACITY_RE = re.compile(
    r"^\d+(?:\.\d+)?(?:TB?|GB|MB)\s+", re.IGNORECASE,
)

# Map category prefix strings to normalized category values
_CATEGORY_PREFIX_MAP: dict[str, str] = {
    "HARD DRIVES": "drive",
    "SSD DRIVES": "ssd",
    "MEMORY": "memory",
    "PROCESSORS": "cpu",
    "NETWORK CARDS": "nic",
    "NETWORK DEVICES": "nic",
    "SWITCHES": "switch",
    "FIREWALLS": "firewall",
    "OPTICS": "optic",
    "GPU": "gpu",
    "MAINBOARD": "server-board",
    "MOTHERBOARD": "server-board",
    "SERVER": "server",
    "SERVER BAREBONE": "server",
    "CHASSIS": "chassis",
    "RAID CARDS": "raid",
    "RAID": "raid",
    "HEAT SINKS": "cooling",
    "COOLING": "cooling",
}


def _detect_category_from_prefix(raw_model: str) -> str | None:
    """Extract category from a CATEGORY:CONDITION: prefix in the raw model string."""
    idx = raw_model.find(":")
    if idx <= 0:
        return None
    prefix = raw_model[:idx].strip().upper()
    return _CATEGORY_PREFIX_MAP.get(prefix)


def _detect_manufacturer(model: str, category: str) -> str | None:
    """Attempt to detect manufacturer from model name and category.

    Returns the manufacturer name if a match is found, None otherwise.
    Strips category/condition/capacity prefixes before matching.
    """
    # Strip category:condition: and capacity prefixes for robust matching
    stripped = _DETECTION_PREFIX_RE.sub("", model).strip()
    stripped = _DETECTION_CAPACITY_RE.sub("", stripped).strip()
    if not stripped:
        return None

    # Check category-specific rules first (higher specificity)
    cat_lower = category.lower()
    for rule_cat, pattern, mfr in _CATEGORY_MANUFACTURER_RULES:
        if cat_lower == rule_cat and pattern.search(stripped):
            return mfr

    # Check general rules
    for pattern, mfr in _MANUFACTURER_RULES:
        if pattern.search(stripped):
            return mfr

    return None


def read_models(path: Path, show_warnings: bool = False) -> list[HardwareModel]:
    """Read hardware models from an Excel spreadsheet.

    Expects a sheet named 'Models' (or the first sheet) with columns:
    Model, Manufacturer, Category, Condition, Original_Item.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Models"] if "Models" in wb.sheetnames else wb.worksheets[0]

    rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows)
    headers = [str(h).strip().lower().replace(" ", "_") for h in raw_headers if h is not None]

    models: list[HardwareModel] = []
    no_mfr_rows: list[tuple[int, str]] = []
    for row_num, row in enumerate(rows, start=2):
        values = dict(zip(headers, row))
        # Model priority: MPN > Model > Item
        model_str = str(values.get("mpn") or "").strip()
        if not model_str:
            model_str = str(values.get("model") or "").strip()
        if not model_str:
            model_str = str(values.get("item") or "").strip()
        original_item = str(values.get("item") or "").strip()
        if not model_str:
            continue

        manufacturer = str(values.get("manufacturer") or "").strip()
        if not manufacturer:
            no_mfr_rows.append((row_num, model_str))

        category = str(values.get("category") or "").strip()
        if not category:
            category = (
                _detect_category_from_prefix(original_item)
                or _detect_category_from_prefix(model_str)
                or "unknown"
            )

        models.append(
            HardwareModel(
                model=normalize_model(model_str, manufacturer),
                manufacturer=manufacturer,
                category=category,
                condition=str(values.get("condition") or "").strip(),
                original_item=(
                    original_item if "item" in values
                    else str(values.get("original_item") or "").strip()
                ),
            )
        )

    wb.close()

    if no_mfr_rows:
        if show_warnings:
            for row_num, model_str in no_mfr_rows:
                logger.warning("Row %d: model '%s' has no manufacturer", row_num, model_str)
        else:
            logger.warning(
                "%d rows have no manufacturer (use --show-warnings to list)",
                len(no_mfr_rows),
            )

    # Auto-detect manufacturers for models that have none
    total_empty = sum(1 for m in models if not m.manufacturer)
    detected_count = 0
    for m in models:
        if m.manufacturer:
            continue
        detected = _detect_manufacturer(m.model, m.category)
        if not detected and m.original_item:
            detected = _detect_manufacturer(m.original_item, m.category)
        if detected:
            m.manufacturer = detected
            # Re-normalize with the detected manufacturer so brand-specific
            # rules (prefix stripping, etc.) are applied.
            m.model = normalize_model(m.model, detected)
            detected_count += 1
            logger.info("Auto-detected manufacturer '%s' for model '%s'", detected, m.model)
    if total_empty:
        logger.info(
            "Auto-detected manufacturers for %d of %d models without manufacturer",
            detected_count,
            total_empty,
        )

    return models


def write_results(
    results: list[EOLResult],
    path: Path,
    filtered_rows: list[dict] | None = None,
) -> None:
    """Write EOL check results to an xlsx file with formatting."""
    wb = openpyxl.Workbook()

    # --- EOL Results sheet ---
    ws = wb.active
    ws.title = "EOL Results"

    # Header row
    for col, name in enumerate(_RESULT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=r.model.model)
        ws.cell(row=row_idx, column=2, value=r.model.manufacturer)
        ws.cell(row=row_idx, column=3, value=r.model.category)
        ws.cell(row=row_idx, column=4, value=r.model.condition)

        status_cell = ws.cell(row=row_idx, column=5, value=r.status.value)
        style = _STATUS_STYLES.get(r.status.value)
        if style:
            status_cell.font = style[0]
            status_cell.fill = style[1]

        ws.cell(row=row_idx, column=6, value=str(r.eol_date) if r.eol_date else "")
        ws.cell(row=row_idx, column=7, value=str(r.eos_date) if r.eos_date else "")
        ws.cell(row=row_idx, column=8, value=str(r.release_date) if r.release_date else "")
        ws.cell(
            row=row_idx, column=9,
            value=_date_source_label(r.date_source),
        )
        ws.cell(row=row_idx, column=10, value=r.confidence)
        ws.cell(row=row_idx, column=11, value=r.source_name)
        ws.cell(row=row_idx, column=12, value=r.checked_at.isoformat())
        ws.cell(row=row_idx, column=13, value=r.model.original_item)
        ws.cell(row=row_idx, column=14, value=r.notes)
        ws.cell(row=row_idx, column=15, value=r.eol_reason.value)

        risk_cell = ws.cell(row=row_idx, column=16, value=r.risk_category.value)
        risk_style = _RISK_STYLES.get(r.risk_category.value)
        if risk_style:
            risk_cell.font = risk_style[0]
            risk_cell.fill = risk_style[1]

    # Column widths
    for col, name in enumerate(_RESULT_COLUMNS, start=1):
        width = min(_DEFAULT_WIDTHS.get(name, 15), 50)
        ws.column_dimensions[get_column_letter(col)].width = width

    # Auto-filter and freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_RESULT_COLUMNS))}{len(results) + 1}"
    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws_summary = wb.create_sheet("Summary")

    ws_summary.cell(row=1, column=1, value="EOL Results Summary")
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws_summary.cell(row=2, column=1, value="Generated At")
    ws_summary.cell(row=2, column=2, value=datetime.now().isoformat(timespec="seconds"))

    ws_summary.cell(row=3, column=1, value="Total Models Checked")
    ws_summary.cell(row=3, column=2, value=len(results))

    # Count by status
    status_counts: dict[str, int] = {}
    for r in results:
        status_counts[r.status.value] = status_counts.get(r.status.value, 0) + 1

    ws_summary.cell(row=4, column=1, value="Total EOL")
    ws_summary.cell(row=4, column=2, value=status_counts.get("eol", 0))
    ws_summary.cell(row=5, column=1, value="Total Active")
    ws_summary.cell(row=5, column=2, value=status_counts.get("active", 0))
    ws_summary.cell(row=6, column=1, value="Total Unknown")
    ws_summary.cell(row=6, column=2, value=status_counts.get("unknown", 0))

    # Count by risk category
    risk_counts: dict[str, int] = {}
    for r in results:
        risk_counts[r.risk_category.value] = risk_counts.get(r.risk_category.value, 0) + 1

    ws_summary.cell(row=7, column=1, value="Risk: Security")
    ws_summary.cell(row=7, column=2, value=risk_counts.get("security", 0))
    ws_summary.cell(row=8, column=1, value="Risk: Support")
    ws_summary.cell(row=8, column=2, value=risk_counts.get("support", 0))
    ws_summary.cell(row=9, column=1, value="Risk: Procurement")
    ws_summary.cell(row=9, column=2, value=risk_counts.get("procurement", 0))
    ws_summary.cell(row=10, column=1, value="Risk: Informational")
    ws_summary.cell(row=10, column=2, value=risk_counts.get("informational", 0))

    # Pivot: status counts by manufacturer
    row_offset = 12
    ws_summary.cell(row=row_offset, column=1, value="EOL Status by Manufacturer")
    ws_summary.cell(row=row_offset, column=1).font = Font(bold=True, size=12)

    mfr_status: dict[str, dict[str, int]] = {}
    for r in results:
        mfr = r.model.manufacturer or "(no manufacturer)"
        mfr_status.setdefault(mfr, {})
        sv = r.status.value
        mfr_status[mfr][sv] = mfr_status[mfr].get(sv, 0) + 1

    statuses = ["eol", "eol_announced", "active", "unknown", "not_found"]
    header_row = row_offset + 1
    ws_summary.cell(row=header_row, column=1, value="Manufacturer")
    ws_summary.cell(row=header_row, column=1).font = Font(bold=True)
    for ci, s in enumerate(statuses, start=2):
        cell = ws_summary.cell(row=header_row, column=ci, value=s)
        cell.font = Font(bold=True)

    for ri, (mfr, counts) in enumerate(sorted(mfr_status.items()), start=header_row + 1):
        ws_summary.cell(row=ri, column=1, value=mfr)
        for ci, s in enumerate(statuses, start=2):
            ws_summary.cell(row=ri, column=ci, value=counts.get(s, 0))

    # Auto-size summary columns
    ws_summary.column_dimensions["A"].width = 25
    for ci in range(2, len(statuses) + 2):
        ws_summary.column_dimensions[get_column_letter(ci)].width = 16

    # --- Filtered sheet (optional) ---
    if filtered_rows:
        ws_filtered = wb.create_sheet("Filtered")
        filtered_headers = ["Model", "Manufacturer", "Reason"]
        for col, name in enumerate(filtered_headers, start=1):
            cell = ws_filtered.cell(row=1, column=col, value=name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        for row_idx, f in enumerate(filtered_rows, start=2):
            ws_filtered.cell(row=row_idx, column=1, value=f["model"])
            ws_filtered.cell(row=row_idx, column=2, value=f["manufacturer"])
            ws_filtered.cell(row=row_idx, column=3, value=f["reason"])
        ws_filtered.column_dimensions["A"].width = 30
        ws_filtered.column_dimensions["B"].width = 20
        ws_filtered.column_dimensions["C"].width = 30

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
