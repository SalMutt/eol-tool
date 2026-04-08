"""FastAPI application wrapping eol-tool functionality."""

import csv
import io
import json
import logging
import re
import threading
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
from fastapi import FastAPI, File, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from fastapi.staticfiles import StaticFiles
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, field_validator

from eol_tool import __version__

from .cache import ResultCache
from .check_pipeline import run_check_pipeline
from .input_filter import filter_models
from .manufacturer_corrections import apply_manufacturer_corrections
from .models import EOLResult, HardwareModel
from .normalizer import normalize_model
from .reader import read_models
from .registry import list_checkers

logger = logging.getLogger(__name__)

_VALID_STATUSES = {"eol", "active", "unknown"}
_VALID_REASONS = {
    "manufacturer_declared", "technology_generation", "product_discontinued",
    "vendor_acquired", "community_data", "manual_override", "none", "",
}
_VALID_RISKS = {"security", "support", "procurement", "informational", "none", ""}
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_CSV_FIELDS = [
    "model", "manufacturer", "status", "eol_reason", "risk_category",
    "eol_date", "eos_date", "source_url", "notes",
]

_CSV_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "manual_overrides.csv"
_LAST_RUN_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "last_run.json"
_csv_lock = threading.Lock()
_last_run_lock = threading.Lock()


class OverrideBody(BaseModel):
    model: str
    manufacturer: str = ""
    status: str
    eol_reason: str = ""
    risk_category: str = ""
    eol_date: str = ""
    eos_date: str = ""
    source_url: str = ""
    notes: str = ""

    @field_validator("model")
    @classmethod
    def model_non_empty(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("model is required and must be non-empty")
        return v.strip()

    @field_validator("status")
    @classmethod
    def status_valid(cls, v: str) -> str:
        if v.lower() not in _VALID_STATUSES:
            raise ValueError(f"status must be one of: {', '.join(sorted(_VALID_STATUSES))}")
        return v.lower()

    @field_validator("eol_reason")
    @classmethod
    def reason_valid(cls, v: str) -> str:
        if v and v.lower() not in _VALID_REASONS:
            valid = ", ".join(sorted(_VALID_REASONS - {""}))
            raise ValueError(f"eol_reason must be one of: {valid}")
        return v.lower()

    @field_validator("risk_category")
    @classmethod
    def risk_valid(cls, v: str) -> str:
        if v and v.lower() not in _VALID_RISKS:
            valid = ", ".join(sorted(_VALID_RISKS - {""}))
            raise ValueError(f"risk_category must be one of: {valid}")
        return v.lower()

    @field_validator("eol_date", "eos_date")
    @classmethod
    def date_format(cls, v: str) -> str:
        if v and not _DATE_RE.match(v):
            raise ValueError("date must be in YYYY-MM-DD format")
        return v


class OverrideDeleteParams(BaseModel):
    model: str
    manufacturer: str = ""


def _read_overrides_csv(csv_path: Path | None = None) -> list[dict[str, str]]:
    path = csv_path or _CSV_PATH
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _write_overrides_csv(rows: list[dict[str, str]], csv_path: Path | None = None) -> None:
    path = csv_path or _CSV_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".csv.tmp")
    with open(tmp, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in _CSV_FIELDS})
    tmp.replace(path)


def _override_key(row: dict[str, str]) -> tuple[str, str]:
    return (row.get("model", "").strip().lower(), row.get("manufacturer", "").strip().lower())


def get_csv_path() -> Path:
    """Return the CSV path. Allows tests to override."""
    return _CSV_PATH


app = FastAPI(title="EOL Tool API", version=__version__)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _result_to_dict(r: EOLResult) -> dict:
    """Convert an EOLResult to the API response dict."""
    return {
        "model": r.model.model,
        "manufacturer": r.model.manufacturer,
        "category": r.model.category,
        "status": r.status.value,
        "eol_date": r.eol_date.isoformat() if r.eol_date else None,
        "eos_date": r.eos_date.isoformat() if r.eos_date else None,
        "date_source": r.date_source,
        "risk_category": r.risk_category.value,
        "eol_reason": r.eol_reason.value,
        "confidence": r.confidence,
        "source": r.source_name,
        "notes": r.notes,
    }


def _infer_manufacturer(model_str: str) -> str:
    """Try to infer manufacturer from model string using known checker prefixes."""
    upper = model_str.upper()
    prefix_map = {
        "EX": "Juniper",
        "QFX": "Juniper",
        "MX": "Juniper",
        "SRX": "Juniper",
        "ACX": "Juniper",
        "JNP": "Juniper",
        "WS-C": "Cisco",
        "N9K": "Cisco",
        "N5K": "Cisco",
        "N3K": "Cisco",
        "C9": "Cisco",
        "ASR": "Cisco",
        "ISR": "Cisco",
        "XEON": "Intel",
        "X520": "Intel",
        "X540": "Intel",
        "X550": "Intel",
        "X710": "Intel",
        "E5-": "Intel",
        "E3-": "Intel",
        "EPYC": "AMD",
        "OPTERON": "AMD",
        "POWEREDGE": "Dell",
        "MZ-": "Samsung",
        "PM": "Samsung",
        "MZIL": "Samsung",
        "ST": "Seagate",
        "X10": "Supermicro",
        "X11": "Supermicro",
        "X12": "Supermicro",
        "X13": "Supermicro",
        "MCX": "Mellanox",
        "CX": "Mellanox",
        "MTFD": "Micron",
        "5300": "Micron",
        "5400": "Micron",
        "WD": "WD",
        "WUS": "WD",
    }
    for prefix, mfr in prefix_map.items():
        if upper.startswith(prefix):
            return mfr
    return ""


@app.get("/api/health")
async def health():
    """Health check endpoint."""
    return {"status": "ok", "version": __version__}


@app.get("/api/lookup")
async def lookup(
    model: str = Query(..., description="Model string to look up"),
    manufacturer: str = Query("", description="Manufacturer name (optional)"),
):
    """Look up EOL status for a single model."""
    if not manufacturer:
        manufacturer = _infer_manufacturer(model)

    normalized = normalize_model(model, manufacturer)
    hw = HardwareModel(
        model=normalized,
        manufacturer=manufacturer,
        category="unknown",
        original_item=model,
    )

    apply_manufacturer_corrections([hw])
    results = await run_check_pipeline([hw])
    if not results:
        return {"model": model, "status": "not_found", "date_source": "none"}

    r = results[0]
    return _result_to_dict(r)


@app.post("/api/check")
async def check_upload(file: UploadFile = File(...)):
    """Upload an xlsx file and run the full check pipeline."""
    suffix = ".xlsx"
    with NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = Path(tmp.name)

    try:
        models = read_models(tmp_path)
        apply_manufacturer_corrections(models)
        models, filtered_rows = filter_models(models)
        results = await run_check_pipeline(models)
    finally:
        tmp_path.unlink(missing_ok=True)

    status_counts: dict[str, int] = {}
    for r in results:
        status_counts[r.status.value] = status_counts.get(r.status.value, 0) + 1

    dated_count = sum(1 for r in results if r.eol_date is not None)

    result_dicts = [_result_to_dict(r) for r in results]

    # Save snapshot for diff comparison
    _save_last_run(result_dicts)

    return {
        "total": len(results),
        "filtered": len(filtered_rows),
        "eol": status_counts.get("eol", 0) + status_counts.get("eol_announced", 0),
        "active": status_counts.get("active", 0),
        "unknown": status_counts.get("unknown", 0),
        "not_found": status_counts.get("not_found", 0),
        "dated": dated_count,
        "results": result_dicts,
    }


@app.get("/api/sources")
async def sources():
    """List available data sources and their cache freshness."""
    cache_inst = ResultCache()
    try:
        source_info = await cache_inst.source_stats()
    finally:
        await cache_inst.close()

    checkers = list_checkers()

    # Build source list from registered checkers + cache info
    source_cache_map = {s["source"]: s for s in source_info}
    source_list = []

    # Known source types
    source_types = {
        "endoflife.date": "api",
        "juniper": "scraper",
        "supermicro": "scraper",
        "cisco": "scraper",
        "dell": "scraper",
    }

    seen = set()
    for name, cls in sorted(checkers.items()):
        if name.startswith("__"):
            if name == "__fallback__":
                display_name = "endoflife.date"
            elif name == "__manual__":
                display_name = "manual-overrides"
            elif name == "__techgen__":
                display_name = "tech-generation"
            else:
                continue
        else:
            display_name = name

        if display_name in seen:
            continue
        seen.add(display_name)

        cache_entry = source_cache_map.get(display_name)
        src_type = source_types.get(display_name, "local")

        entry: dict = {
            "name": display_name,
            "type": src_type,
            "priority": cls.priority,
        }

        if cache_entry:
            entry["cached_products"] = cache_entry["item_count"]
            fetched_at = cache_entry["fetched_at"]
            if fetched_at:
                age_hours = (datetime.now() - fetched_at).total_seconds() / 3600
                entry["cache_age_hours"] = round(age_hours, 1)
                entry["last_refreshed"] = fetched_at.isoformat() + "Z"
            else:
                entry["cache_age_hours"] = None
                entry["last_refreshed"] = None
        else:
            entry["cached_products"] = 0
            entry["cache_age_hours"] = None
            entry["last_refreshed"] = None

        source_list.append(entry)

    return {"sources": source_list}


@app.get("/api/overrides")
async def list_overrides():
    """Return all manual overrides as a JSON array."""
    with _csv_lock:
        rows = _read_overrides_csv(get_csv_path())
    return [
        {k: row.get(k, "") for k in _CSV_FIELDS}
        for row in rows
        if row.get("model", "").strip()
    ]


@app.post("/api/overrides", status_code=201)
async def create_override(body: OverrideBody):
    """Add a new manual override."""
    new_key = (body.model.lower(), body.manufacturer.strip().lower())
    with _csv_lock:
        rows = _read_overrides_csv(get_csv_path())
        for row in rows:
            if _override_key(row) == new_key:
                return Response(
                    content='{"detail":"duplicate model+manufacturer combination"}',
                    status_code=409,
                    media_type="application/json",
                )
        new_row = body.model_dump()
        rows.append(new_row)
        _write_overrides_csv(rows, get_csv_path())
    return {k: new_row.get(k, "") for k in _CSV_FIELDS}


@app.put("/api/overrides")
async def update_override(body: OverrideBody):
    """Update an existing manual override identified by model+manufacturer."""
    target_key = (body.model.lower(), body.manufacturer.strip().lower())
    with _csv_lock:
        rows = _read_overrides_csv(get_csv_path())
        found = False
        for i, row in enumerate(rows):
            if _override_key(row) == target_key:
                rows[i] = body.model_dump()
                found = True
                break
        if not found:
            return Response(
                content='{"detail":"override not found"}',
                status_code=404,
                media_type="application/json",
            )
        _write_overrides_csv(rows, get_csv_path())
    return {k: rows[i].get(k, "") for k in _CSV_FIELDS}


@app.delete("/api/overrides")
async def delete_override(
    model: str = Query(..., description="Model to delete"),
    manufacturer: str = Query("", description="Manufacturer"),
):
    """Delete a manual override by model+manufacturer."""
    target_key = (model.strip().lower(), manufacturer.strip().lower())
    with _csv_lock:
        rows = _read_overrides_csv(get_csv_path())
        new_rows = [r for r in rows if _override_key(r) != target_key]
        if len(new_rows) == len(rows):
            return Response(
                content='{"detail":"override not found"}',
                status_code=404,
                media_type="application/json",
            )
        _write_overrides_csv(new_rows, get_csv_path())
    return {"deleted": True}


@app.get("/api/overrides/export")
async def export_overrides():
    """Download the overrides CSV file."""
    with _csv_lock:
        rows = _read_overrides_csv(get_csv_path())
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_CSV_FIELDS)
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in _CSV_FIELDS})
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=manual_overrides.csv"},
    )


@app.post("/api/overrides/import")
async def import_overrides(file: UploadFile = File(...)):
    """Import a CSV file and merge with existing overrides."""
    content = (await file.read()).decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    incoming = list(reader)

    with _csv_lock:
        existing = _read_overrides_csv(get_csv_path())
        existing_map: dict[tuple[str, str], int] = {}
        for i, row in enumerate(existing):
            existing_map[_override_key(row)] = i

        added = 0
        updated = 0
        unchanged = 0

        for imp_row in incoming:
            key = _override_key(imp_row)
            if not key[0]:
                continue
            normalized = {k: imp_row.get(k, "") for k in _CSV_FIELDS}
            if key in existing_map:
                idx = existing_map[key]
                old = {k: existing[idx].get(k, "") for k in _CSV_FIELDS}
                if old != normalized:
                    existing[idx] = normalized
                    updated += 1
                else:
                    unchanged += 1
            else:
                existing.append(normalized)
                existing_map[key] = len(existing) - 1
                added += 1

        _write_overrides_csv(existing, get_csv_path())

    return {"added": added, "updated": updated, "unchanged": unchanged}


def _save_last_run(result_dicts: list[dict]) -> None:
    """Save a snapshot of the current run for diff comparison.

    Rotates the previous last_run.json to prev_run.json before writing.
    """
    snapshot = {
        "timestamp": datetime.now().isoformat(),
        "results": result_dicts,
    }
    with _last_run_lock:
        _LAST_RUN_PATH.parent.mkdir(parents=True, exist_ok=True)
        # Rotate: current last_run -> prev_run
        prev_path = _LAST_RUN_PATH.with_name("prev_run.json")
        if _LAST_RUN_PATH.exists():
            _LAST_RUN_PATH.replace(prev_path)
        # Write new last_run
        tmp = _LAST_RUN_PATH.with_suffix(".json.tmp")
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(snapshot, fh)
        tmp.replace(_LAST_RUN_PATH)


def _load_last_run() -> dict | None:
    """Load the previous run snapshot, or None if not available."""
    with _last_run_lock:
        if not _LAST_RUN_PATH.exists():
            return None
        with open(_LAST_RUN_PATH, encoding="utf-8") as fh:
            return json.load(fh)


@app.get("/api/diff")
async def diff():
    """Compare current results against the previous run snapshot."""
    snapshot = _load_last_run()
    if not snapshot:
        return {
            "message": "No previous run data available. Upload a file via /api/check first.",
            "changes": [],
            "summary": {"new_eol": 0, "new_models": 0, "status_changes": 0},
        }

    # Load current cached results (re-read the snapshot as "current")
    # The last_run.json IS the most recent run. To compute a diff we need
    # both the current and the *previous* run. We keep two files:
    # last_run.json (current) and prev_run.json (previous).
    prev_path = _LAST_RUN_PATH.with_name("prev_run.json")
    with _last_run_lock:
        if not prev_path.exists():
            return {
                "message": "Only one run recorded so far. Run another check to see changes.",
                "changes": [],
                "summary": {"new_eol": 0, "new_models": 0, "status_changes": 0},
                "current_run": snapshot.get("timestamp"),
            }
        with open(prev_path, encoding="utf-8") as fh:
            prev_snapshot = json.load(fh)

    prev_results = prev_snapshot.get("results", [])
    curr_results = snapshot.get("results", [])

    # Build lookup by (model, manufacturer)
    prev_map: dict[tuple[str, str], dict] = {}
    for r in prev_results:
        key = (r.get("model", "").lower(), r.get("manufacturer", "").lower())
        prev_map[key] = r

    curr_map: dict[tuple[str, str], dict] = {}
    for r in curr_results:
        key = (r.get("model", "").lower(), r.get("manufacturer", "").lower())
        curr_map[key] = r

    changes = []

    # New models and status changes
    for key, curr in curr_map.items():
        prev = prev_map.get(key)
        if prev is None:
            changes.append({
                "model": curr.get("model", ""),
                "manufacturer": curr.get("manufacturer", ""),
                "previous_status": None,
                "current_status": curr.get("status", ""),
                "change_type": "new_model",
            })
        elif prev.get("status") != curr.get("status"):
            change_type = "status_change"
            if curr.get("status") == "eol" and prev.get("status") in ("active", "unknown"):
                change_type = "new_eol"
            changes.append({
                "model": curr.get("model", ""),
                "manufacturer": curr.get("manufacturer", ""),
                "previous_status": prev.get("status", ""),
                "current_status": curr.get("status", ""),
                "change_type": change_type,
            })

    # Removed models
    for key, prev in prev_map.items():
        if key not in curr_map:
            changes.append({
                "model": prev.get("model", ""),
                "manufacturer": prev.get("manufacturer", ""),
                "previous_status": prev.get("status", ""),
                "current_status": None,
                "change_type": "removed",
            })

    summary = {
        "new_eol": sum(1 for c in changes if c["change_type"] == "new_eol"),
        "new_models": sum(1 for c in changes if c["change_type"] == "new_model"),
        "status_changes": sum(1 for c in changes if c["change_type"] == "status_change"),
        "removed": sum(1 for c in changes if c["change_type"] == "removed"),
    }

    return {
        "changes": changes,
        "summary": summary,
        "current_run": snapshot.get("timestamp"),
        "previous_run": prev_snapshot.get("timestamp"),
    }


_EXPORT_COLUMNS = [
    "Model", "Manufacturer", "Category", "EOL Status", "EOL Date", "EOS Date",
    "Date Source", "Risk Category", "EOL Reason", "Confidence", "Source", "Notes",
]

_EXPORT_WIDTHS = {
    "Model": 35, "Manufacturer": 18, "Category": 14, "EOL Status": 14,
    "EOL Date": 14, "EOS Date": 14, "Date Source": 22, "Risk Category": 16,
    "EOL Reason": 22, "Confidence": 12, "Source": 25, "Notes": 40,
}

_EXPORT_STATUS_STYLES = {
    "eol": (Font(color="FF0000"), PatternFill(fgColor="FFE6E6", fill_type="solid")),
    "eol_announced": (Font(color="FF8C00"), PatternFill(fgColor="FFF3E0", fill_type="solid")),
    "active": (Font(color="008000"), PatternFill(fgColor="E6FFE6", fill_type="solid")),
}

_EXPORT_RISK_STYLES = {
    "security": (Font(color="8B0000"), PatternFill(fgColor="FFE0E0", fill_type="solid")),
    "support": (Font(color="FF8C00"), PatternFill(fgColor="FFF0E0", fill_type="solid")),
    "procurement": (Font(color="0000CD"), PatternFill(fgColor="E0E8FF", fill_type="solid")),
}

_EXPORT_HEADER_FONT = Font(bold=True, color="FFFFFF")
_EXPORT_HEADER_FILL = PatternFill(fgColor="2F5496", fill_type="solid")

_DATE_SOURCE_LABELS = {
    "manufacturer_confirmed": "Manufacturer Confirmed",
    "community_database": "Community Database",
}

_EOL_REASON_LABELS = {
    "manufacturer_declared": "Manufacturer Declared",
    "technology_generation": "Technology Generation",
    "product_discontinued": "Product Discontinued",
    "vendor_acquired": "Vendor Acquired",
    "community_data": "Community Data",
    "manual_override": "Manual Override",
}


@app.get("/api/export")
async def export_results(
    risk_category: str = Query("", description="Filter by risk category"),
    status: str = Query("", description="Filter by EOL status"),
):
    """Export the last check results as an xlsx file, with optional filters."""
    snapshot = _load_last_run()
    if not snapshot:
        return Response(
            content='{"detail":"No results available. Run a check first."}',
            status_code=404,
            media_type="application/json",
        )

    results = snapshot.get("results", [])

    # Apply filters
    if risk_category:
        results = [r for r in results if r.get("risk_category", "") == risk_category.lower()]
    if status:
        results = [r for r in results if r.get("status", "") == status.lower()]

    # Build xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EOL Report"

    # Header
    for col, name in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _EXPORT_HEADER_FONT
        cell.fill = _EXPORT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=r.get("model", ""))
        ws.cell(row=row_idx, column=2, value=r.get("manufacturer", ""))
        ws.cell(row=row_idx, column=3, value=r.get("category", ""))

        status_val = r.get("status", "")
        status_cell = ws.cell(row=row_idx, column=4, value=status_val)
        style = _EXPORT_STATUS_STYLES.get(status_val)
        if style:
            status_cell.font = style[0]
            status_cell.fill = style[1]

        ws.cell(row=row_idx, column=5, value=r.get("eol_date") or "")
        ws.cell(row=row_idx, column=6, value=r.get("eos_date") or "")
        ws.cell(
            row=row_idx, column=7,
            value=_DATE_SOURCE_LABELS.get(r.get("date_source", ""), "Not Available"),
        )

        risk_val = r.get("risk_category", "")
        risk_cell = ws.cell(row=row_idx, column=8, value=risk_val)
        risk_style = _EXPORT_RISK_STYLES.get(risk_val)
        if risk_style:
            risk_cell.font = risk_style[0]
            risk_cell.fill = risk_style[1]

        ws.cell(
            row=row_idx, column=9,
            value=_EOL_REASON_LABELS.get(r.get("eol_reason", ""), ""),
        )
        ws.cell(row=row_idx, column=10, value=r.get("confidence", 0))
        ws.cell(row=row_idx, column=11, value=r.get("source", ""))
        ws.cell(row=row_idx, column=12, value=r.get("notes", ""))

    # Column widths
    for col, name in enumerate(_EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = _EXPORT_WIDTHS.get(name, 15)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_EXPORT_COLUMNS))}{len(results) + 1}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    # Build filename with filters
    parts = ["eol-report"]
    if risk_category:
        parts.append(risk_category.lower())
    if status:
        parts.append(status.lower())
    parts.append(datetime.now().strftime("%Y-%m-%d"))
    filename = "-".join(parts) + ".xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Mount static files last so API routes take precedence
_frontend_candidates = [
    Path(__file__).resolve().parent.parent.parent / "frontend",  # source tree
    Path("/app/frontend"),  # Docker
    Path.cwd() / "frontend",  # current directory
]
for _candidate in _frontend_candidates:
    if _candidate.is_dir() and (_candidate / "index.html").is_file():
        app.mount("/", StaticFiles(directory=str(_candidate), html=True), name="frontend")
        break
